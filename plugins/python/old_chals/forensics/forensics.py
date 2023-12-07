import os
from os.path import *
from zipfile import ZipFile
import tempfile
import random
from fractions import Fraction
import piexif
from piexif import GPSIFD
from pdfrw import PdfReader, PdfWriter, PageMerge
from pydub import AudioSegment
from pydub.generators import Sine
from chals.chalenv import ChallengeEnvironment
import pyzipper
import openpyxl
from ..utils import place_letter_highlighted, place_letter, random_alphabet_letter, WorkDir

from chals.utils import text_to_wav

from ..text_transforms import binary_text
from ..text_transforms import text_transforms
from ..challenge import GeneratedChallenge
from ..imgencode import *
from ..docker_builder import DockerBuilder
from stegano import lsb
from shutil import make_archive, copy

TEMPLATES_DIR = join(dirname(dirname(realpath(__file__))), 'templates')


class FileEvidence(ChallengeEnvironment):
    yaml_tag = u'!file_evidence'
    __doc__ = """
    File Evidence

    Config:

        disable_set_file: Set the chal_file to this file
        file: File to include in the challenge
    """

    def gen(self, chal_dir):
        file = self.get_value("file")
        if self.get_value("disable_set_file", required=False):
            return
        self.chal_file = file
        self.chal_host.add_chal(join(chal_dir, file))

class URLEvidence(ChallengeEnvironment):
    yaml_tag = u'!url_evidence'
    __doc__ = """
    File Evidence

    Config:

        url: Url of challenge file
    """

    def gen(self, chal_dir):
        self.display = self.get_value("url")


# class BlankTextFile(GeneratedChallenge):
#     yaml_tag = u'!blank_text_file'
#     __doc__ = """
#     Blank Text File

#     Config:
#         None
#     """

#     def gen(self, chal_dir):
#         binary_flag = binary_text(self.flag)

#         with WorkDir(chal_dir):
#             with open("empty.txt", "w", encoding="utf-8") as f:
#                 for character in binary_flag:
#                     if character == '0':
#                         f.write(" ")
#                     elif character == '1':
#                         f.write(" ")

class LockedExcel(GeneratedChallenge):
    yaml_tag = u'!locked_excel'
    __doc__ = """
    Locked Excel
    
    Config:
        password: Password to lock the excel file with (SHOULD NOT BE REVEALED DURING THE CHALLENGE! Competitors are to break into the excel file WITHOUT a password!)
    """

    def gen(self, chal_dir):
        with WorkDir(chal_dir):
            workbook = openpyxl.Workbook()

            flag = self.flag

            workbook.active.title = "Sheet1"
            workbook.create_sheet(title="Sheet2")

            sheet1 = workbook["Sheet1"]
            sheet2 = workbook["Sheet2"]

            for idx, letter in enumerate(flag):
                row = random.randint(1, 10)
                column = idx + 1
                place_letter_highlighted(sheet1.cell(row=row, column=column), letter)

                iterations = random.randint(1, 5)
                for _ in range(iterations):
                    random_row = random.randint(1, 10)
                    random_column = random.randint(1, len(flag))
                    cell = sheet1.cell(row=random_row, column=random_column)
                    if not cell.value:
                        place_letter(cell, random_alphabet_letter())

            sheet2["A1"] = "???"
            sheet1.sheet_state = "hidden"

            workbook.security.workbookPassword = self.get_value("password")
            workbook.security.lockStructure = True

            workbook.save("locked_workbook.xlsx")

class LSB(GeneratedChallenge):
    yaml_tag = u'!lsb'
    __doc__ = """
    Least Significant Bit Encoding

    Config:

        orig_img - The relative path of the image to use
    """

    def gen(self, chal_dir):
        filename = self.get_value("orig_img")
        out = lsb.hide(os.path.join(chal_dir, filename),
                       self.flag + self.get_value("addl_text"))
        out.save(os.path.join(chal_dir, "chal-img.png"))
        self.chal_file = "chal-img.png"


class DiffImages(GeneratedChallenge):
    yaml_tag = u'!diffimages'
    __doc__ = """
    Flag is located in the byte difference between two images

    Config:

        orig_img - The relative path of the image to use
    """

    def gen(self, chal_dir):
        filename = self.get_value("orig_img")
        orig_img_path = os.path.join(chal_dir, filename)
        with open(orig_img_path, 'rb') as f:
            raw_data = f.read()
            bytelist = bytearray(raw_data)

        def insertFlag(mssg):

            random_offset = random.randint(24, int(len(bytelist) / 4))
            available_bytes = int((len(bytelist) - 4 - random_offset) / len(mssg))
            for i, c in enumerate(mssg):
                random_num = random.randint(random_offset + (i * available_bytes), random_offset + ((i + 1) * available_bytes))
                bytelist[random_num:random_num+1] = c.encode('utf-8')

            return bytelist
        flag = self.flag + self.get_value("addl_text")
        insertFlag(flag)

        outputfile = 'chal.png'
        output_img_path = os.path.join(chal_dir, outputfile)
        with open(output_img_path, 'wb') as image_file:
            image_file.write(bytelist)

        orig_filename = '_' + os.path.basename(filename)
        new_orig_img_path = os.path.join(chal_dir, orig_filename)
        copy(orig_img_path, new_orig_img_path)

        self.chal_file = [orig_filename, outputfile]


class BrokenPNGHeaderChallenge(GeneratedChallenge):
    yaml_tag = u'!brokenpngheader'
    __doc__ = """
    Flag is located in the image, whose header is broken

    Config:

        orig_img - The relative path of the image to use
    """

    def gen(self, chal_dir):
        orig_img = self.get_value("orig_img")
        with open(os.path.join(chal_dir, orig_img), 'rb') as f:
            content = f.read()
            bytelist = bytearray(content)
        c = "c"
        bytelist[0:1] = c.encode('utf-8')
        outputfile = 'edited_photo.png'
        with open(os.path.join(chal_dir, outputfile), 'wb') as image_file:
            image_file.write(bytelist)
        self.chal_file = outputfile

    def solve(self, chal_dir):
        pass


class SpectrogramChallenge(GeneratedChallenge):
    yaml_tag = u'!spectrogram'
    __doc__ = """
    Flag is located in the spectrum data of the audio file
    
    Config:

        orig_img - The relative path of the image to use
        output_wav - The relative path to the wav output file
    """

    def gen(self, chal_dir):
        orig_img = self.get_value("orig_img")
        output_wav = self.get_value("output_wav")
        flag = self.flag
        seconds = int(len(flag)/3)
        start(os.path.join(chal_dir, orig_img),
              os.path.join(chal_dir, output_wav), seconds)
        self.chal_file = output_wav


class ImageExifChallenge(GeneratedChallenge):
    yaml_tag = u'!image_exif'
    __doc__ = """
    Flag is located in image meta data

    Config:

        orig_img - The relative path of the original image to use
        out_img - The relative path of the output
        metadata_key - The metadata key to store the flag under
        flag_transform - The method to use to transform the flag (so you can't solve it with strings)
        lat - latitude of gps exif
        lng - longitude of gps exif
    """

    def gen(self, chal_dir):
        orig_img = self.get_value("orig_img")
        addl_text = self.get_value("addl_text", False)
        out_img = self.get_value("out_img")
        metadata_key = self.get_value("metadata_key")
        flag_transform = self.get_value("flag_transform", False)
        lat = self.get_value("lat")
        lng = self.get_value("lng")

        self.chal_file = out_img 

        if flag_transform is not None:
            trans_func = text_transforms[flag_transform]["encode"]
        else:
            def trans_func(x): return x

        flag = self.flag
        if addl_text:
            flag += addl_text

        img_path = os.path.join(chal_dir, orig_img)
        img = Image.open(img_path)

        exif_dict = {}
        if 'exif' in img.info.keys() or 'Exif' in img.info.keys():
            exif_dict = piexif.load(img.info['exif'])
        else:
            exif_dict = {
                '0th': {
                    258: (24, 24, 24),
                    271: b'Make',
                    272: b'XXX-XXX',
                    282: (4294967295, 1),
                    296: 65535,
                    305: b'PIL',
                    514: 4294967295,
                    50715: ((1, 1), (1, 1), (1, 1)),
                    34665: 185, 34853: 366
                },
                'Exif': {
                    33434: (4294967295, 1),
                    34856: b'\xaa\xaa\xaa\xaa\xaa\xaa',
                    34867: 4294967295,
                    36867: b'2099:09:29 10:10:10',
                    37380: (2147483647, -2147483648),
                    41994: 65535,
                    42034: ((1, 1), (1, 1), (1, 1), (1, 1)),
                    42035: b'LensMake'
                },
                'GPS': {
                    0: (0, 0, 0, 1),
                    2: (4294967295, 1),
                    5: 1,
                    29: b'1999:99:99 99:99:99',
                    30: 65535
                },
                'Interop': {},
                '1st': {},
                'thumbnail': None
            }
            piexif.insert(piexif.dump(exif_dict), img_path)

        def dec_to_deg(dec):
            positive = dec >= 0
            dec = abs(dec)

            degrees = int(dec)
            minutes = (dec - int(dec)) * 60
            seconds = (minutes - int(minutes)) * 60
            return tuple([
                Fraction(num).limit_denominator(100).as_integer_ratio()
                for num in [degrees, minutes, seconds]
            ]), positive

        deg_lat, ref_lat = dec_to_deg(lat)
        deg_lng, ref_lng = dec_to_deg(lng)
        GPS_IFD = {
            GPSIFD.GPSVersionID: (0, 0, 0, 1),  # byte
            GPSIFD.GPSLatitudeRef: 'E' if ref_lat else 'W',
            GPSIFD.GPSLatitude: deg_lat,  # rational
            GPSIFD.GPSLongitudeRef: 'N' if ref_lng else 'S',
            GPSIFD.GPSLongitude: deg_lng,  # rational
        }
        exif_dict['GPS'] = GPS_IFD

        exif_dict['0th'][metadata_key] = trans_func(flag)
        exif_dict['Exif'][41729] = b'1'
        exif_bytes = piexif.dump(exif_dict)

        # TODO Support different image types
        img_out_path = os.path.join(chal_dir, self.chal_file)
        img.save(img_out_path, "jpeg", exif=exif_bytes)

    def solve(self, chal_dir):
        orig_img = self.get_value("orig_img")
        metadata_key = self.get_value("metadata_key")
        flag_transform = self.get_value("flag_transform")
        trans_func = text_transforms[flag_transform]["decode"]

        img = Image.open(os.path.join(chal_dir, orig_img))

        exif_dict = {}
        if 'exif' in img.info.keys():
            exif_dict = piexif.load(img.info['exif'])

        exif_dict[metadata_key] = trans_func(self.flag)
        exif_bytes = piexif.dump(exif_dict)

        # TODO: Support different image types
        img_out_path = os.path.join(chal_dir, 'chal_img.jpeg')
        img.save(img_out_path, "jpeg", exif=exif_bytes)

class GitReassemblyChallenge(GeneratedChallenge):
    yaml_tag = u'!git_reassembly'
    __doc__ = """
    Determine flag by reassambling parts from git

    Config:

        text - extra text to include in git file
        num_parts - number of parts to split the flag into
        git_email - email of user 
        git_name - name of user
    """

    def gen(self, chal_dir):
        text = self.get_value("text")
        num_parts = self.get_value("num_parts")

        git_email = "test@test.com"
        git_name = "Test"
        configured_git_email = self.get_value("git_email", required=False)
        configured_git_name = self.get_value("git_name", required=False)
        git_email = configured_git_email if configured_git_email is not None else git_email
        git_name = configured_git_name if configured_git_name is not None else git_name

        full_text = text + ": " + self.flag
        git_folder = "/git_chal"
        input_dir = "input"

        with tempfile.TemporaryDirectory() as chal_context:
            def write_tmp(name, contents):
                path = join(chal_context, name)
                with open(path, "w") as f:
                    f.write(contents)
                return path

            included_files = []
            for i in range(0, len(full_text), 3):
                if i + 3 < len(full_text) - 1:
                    part = full_text[i:i+3]
                else:
                    part = full_text[i:]
                filename = f"flag_{i}.txt"
                included_file_path = write_tmp(filename, part)
                included_files.append(included_file_path)

            builder = DockerBuilder(
                base_img="bitnami/git",
                input_dir=input_dir,
                included_files=included_files,
                output_files=[join("/", git_folder)],
                outdir=chal_context
            )

            with builder as b:
                b.run("git config --global user.email \"test@test.com\"")
                b.run("git config --global user.name \"Test\"")
                b.run(f"mkdir {git_folder}")
                b.run(f"git init {git_folder}")
                b.run(f"cd {git_folder}")
                for included_file in included_files:
                    docker_included_file = join(
                        "..", input_dir, os.path.basename(included_file))
                    git_flag_file = join(git_folder, "flag.txt")
                    b.run(f"cp {docker_included_file} {git_flag_file}")
                    b.run("git add flag.txt")
                    b.run("git commit -m 'Updating flag.txt'")
                    b.run("rm flag.txt")
                b.run("git rm flag.txt")
                b.run("git commit -m 'Removing flag.txt'")

            chal_zip = join(chal_dir, "chal")
            outdir = join(chal_context, 'git_chal')
            make_archive(chal_zip, "zip", outdir)
            self.chal_file = "chal.zip"

    def solve(self, chal_dir):
        pass


class GitRevertChallenge(GeneratedChallenge):
    yaml_tag = u'!git_revert'
    __doc__ = """
    Determine secret information from old commits

    Config:
    
        text - extra text to include in git file
        file - non-secret file name
        file_contents - non-secret file contents
        git_email - email of user 
        git_name - name of user
    """

    def gen(self, chal_dir):
        text = self.get_value("text")

        configured_git_email = self.get_value("git_email", required=False)
        configured_git_name = self.get_value("git_name", required=False)
        git_email = configured_git_email if configured_git_email is not None else "test@test.com"
        git_name = configured_git_name if configured_git_name is not None else "Test"
        configured_file = self.get_value("file", required=False)
        configured_contents = self.get_value("file_contents", required=False)
        file_name = configured_file if configured_file is not None else "message.txt"
        file_contents = configured_contents if configured_contents is not None else 'nothing to see here ;)'

        full_text = text + ": " + self.flag
        git_folder = "git_chal"

        with tempfile.TemporaryDirectory() as chal_context:
            builder = DockerBuilder(
                base_img="bitnami/git",
                output_files=[join("/", git_folder)],
                outdir=chal_context
            )

            with builder as b:
                b.run(f"git config --global user.email \"{git_email}\"")
                b.run(f"git config --global user.name \"{git_name}\"")
                b.run(f"mkdir {git_folder}")
                b.run(f"git init {git_folder}")
                b.run(f"cd {git_folder}")
                b.run(f"echo '{full_text}' > flag.txt")
                b.run(f"echo '{file_contents}' > {file_name}")
                b.run("git add flag.txt")
                b.run(f"git add {file_name}")
                b.run("git commit -m 'Initial commit'")
                b.run("rm flag.txt")
                b.run(f"git add flag.txt")
                b.run("git commit -m 'Removing sensitive information'")

            chal_zip = os.path.join(chal_dir, "chal")
            outdir = join(chal_context, git_folder)
            make_archive(chal_zip, "zip", outdir)
            self.chal_file = "chal.zip"

    def solve(self, chal_dir):
        pass

def zipdir(path, zipfilename, password):
    """
    Compresses the contents of a directory into a zip archive using a password to encrypt it.

    Args:
    path (str): Path to the directory to be compressed.
    zipfilename (str): Name of the resulting zip archive file.
    password (str): Password to encrypt the zip archive with.
    """
    with pyzipper.AESZipFile(zipfilename,
                         'w',
                         compression=pyzipper.ZIP_LZMA,
                         encryption=pyzipper.WZ_AES) as zipf:
        # Set a password to encrypt the zip archive
        zipf.setpassword(password.encode())
        # Recursively compress each file in the directory
        for root, dirs, files in os.walk(path):
            for file in files:
                zipf.write(os.path.join(root, file), arcname=file)

class EncryptedZipChallenge(GeneratedChallenge):
    yaml_tag = u'!encrypted_zip'
    __doc__ = """
    An encrypted zip challenge

    Config:

        files_dir - files to be included with the challenge
        password - the password used to encrypt the files
        word_list - [optional] 
    """

    def gen(self, chal_dir):
        files_dir = self.get_value("files_dir")
        password = self.get_value("password")
        filename = self.get_value("filename", required=False)
        word_list = self.get_value("word_list", required=False)

        self.chal_file = "challenge.zip" if filename is None else filename

        # If we have a wordlist, read it and make sure the configured
        # password exists within it
        words = [password]
        if word_list is not None:
            with open(join(chal_dir, word_list), "r") as w:
                words = [l.strip() for l in w.readlines()]

        assert password in words
        files = os.path.join(chal_dir, files_dir)

        output_zip = os.path.join(chal_dir, self.chal_file)
        zipdir(files, output_zip, password)

    def solve(self, chal_dir):
        pass


class FileCarving(GeneratedChallenge):
    yaml_tag = u'!file_carving'
    __doc__ = """
    Carve out a file from another file

    Config:

        main_file - main file to be released
        hidden_file - file to be hidden in main_file (Note: the contents of this file contain the flag)
    """

    def gen(self, chal_dir):
        main_file = self.get_value("main_file")
        hidden_file = self.get_value("hidden_file")

        main_file_path = os.path.join(chal_dir, main_file)
        hidden_file_path = os.path.join(chal_dir, hidden_file)
        output_file_path = os.path.join(chal_dir, "challenge")

        main_fd = open(main_file_path)
        hidden_fd = open(hidden_file_path)
        output_fd = open(output_file_path, "w")

        output_fd.write(main_fd.read() + hidden_fd.read())

        main_fd.close()
        hidden_fd.close()
        output_fd.close()

    def solve(self, chal_dir):
        pass


class DocxCarving(GeneratedChallenge):
    yaml_tag = u'!docx_carving'
    __doc__ = """
    Carve out files from a docx file

    Config:

        main_file - main file to be released
        hidden_file - file to be hidden in main_file (Note: the contents of this file contain the flag)
        dest_path - path within the docx to put the hidden_file
    """

    def gen(self, chal_dir):
        main_file = self.get_value("main_file")
        out_file = self.get_value("out_file")
        hidden_file = self.get_value("hidden_file")
        dest_path = self.get_value("dest_path")

        self.chal_file = out_file

        main_file_path = os.path.join(chal_dir, main_file)
        hidden_file_path = os.path.join(chal_dir, hidden_file)
        output_file_path = os.path.join(chal_dir, self.chal_file)

        td = tempfile.mkdtemp()
        with ZipFile(main_file_path, "r") as zf:
            zf.extractall(path=td)

        os.system("cp '{}' '{}'".format(
            hidden_file_path, os.path.join(td, dest_path)))

        make_archive(output_file_path, 'zip', td)

        os.system("mv '{}.zip' '{}'".format(
            output_file_path, output_file_path))

    def solve(self, chal_dir):
        pass


class Ext4FileRecovery(GeneratedChallenge):
    yaml_tag = u'!ext4_file_recovery'
    __doc__ = """
    Recover a file 

    Config:

        deleted_evidence - A directory containing the evidence that will be deleted
        fluff - A directory containing data to be added to the filesystem
    """

    def gen(self, chal_dir):
        deleted_evidence = self.get_value("deleted_evidence")
        fluff = self.get_value("fluff")

        deleted_evidence_path = os.path.join(chal_dir, deleted_evidence)
        fluff_path = os.path.join(chal_dir, fluff)
        image_out_path = 'chal.img'
        mount_path = '/chal'

        input_dir = "input"

        included_files = [
            deleted_evidence_path,
            fluff_path
        ]

        outdir = join(chal_dir, image_out_path)
        builder = DockerBuilder(
            base_img="archlinux",
            input_dir=input_dir,
            included_files=included_files,
            output_files=[image_out_path],
            outdir=outdir
        )
        with builder as b:
            b.run(f'dd if=/dev/zero of={image_out_path} bs=1M count=5')
            b.run(f'mkfs.ext4 {image_out_path}')
            b.run(f'mkdir -p {mount_path}')
            b.run(f'mount -t ext4 -o loop,rw {image_out_path} {mount_path}')
            b.run(f'cp -r {deleted_evidence_path}/* {mount_path}')
            b.run(f'rm -rf {mount_path}/*')
            b.run(f'cp -r {fluff_path}/* {mount_path}')
            b.run(f'umount {mount_path}')

        self.chal_file = image_out_path

    def solve(self, chal_dir):
        pass


def get4(srcpages):
    scale = 0.5
    srcpages = PageMerge() + srcpages
    x_increment, y_increment = (scale * i for i in srcpages.xobj_box[2:])
    for i, page in enumerate(srcpages):
        page.scale(scale)
        page.x = x_increment if i & 1 else 0
        page.y = 0 if i & 2 else y_increment
    return srcpages.render()


def scale_pdf(path, output):
    pages = PdfReader(path).pages
    writer = PdfWriter(output)
    scaled_pages = 4

    for i in range(0, len(pages), scaled_pages):
        four_pages = get4(pages[i: i + 4])
        writer.addpage(four_pages)

    writer.write()


class HiddenStuffInPDF(GeneratedChallenge):
    yaml_tag = u'!hidden_stuff_in_pdf'
    __doc__ = """
    Find the hidden stuff in the pdf

    Config:

        file - pdf file to mess with
    """

    def gen(self, chal_dir):
        pdf_file = os.path.join(chal_dir, self.get_value("file"))
        out_file = os.path.join(chal_dir, 'out.pdf')

        # TODO replace this with what we actually want to do, this is just POC code for modifying PDFs
        scale_pdf(pdf_file, out_file)

    def solve(self, chal_dir):
        pass


class ChromiumHistory(GeneratedChallenge):
    yaml_tag = u'!chromium_history'
    __doc__ = """
    Do forensic analysis on a chromium cache folder

    Config:
    
        script - puppeteer script to run 
    """

    def gen(self, chal_dir):
        script = self.get_value("script")

        template_dir = join(TEMPLATES_DIR, 'chrome_history')
        template_script = join(template_dir, 'index.js')

        self.chal_file = 'chrome.tar.gz'
        if script is None:
            script = template_script
        else:
            script = join(chal_dir, script)

        puppet = DockerBuilder(
            base_img='alekzonder/puppeteer',
            input_dir='app',
            included_files=[script],
            output_files=[join('/', 'app', self.chal_file)],
            outdir=chal_dir
        )
        with puppet as p:
            p.run('mkdir history')
            p.run('node index.js', timeout=90)
            p.run(f'tar -cvzf {self.chal_file} history')

    def solve(self, chal_dir):
        pass


class EditContrast(GeneratedChallenge):
    yaml_tag = u'!edit_contrast'
    """Placeholder
    """

    def gen(self, chal_dir):
        pass


class ChangeExtension(GeneratedChallenge):
    yaml_tag = u'!change_extension'
    __doc__ = """
    Change the file extension to access the file

    Config:

        orig_file - txt file
        out_ext - new extension 
    """

    def gen(self, chal_dir):
        orig_img = self.get_value("orig_file")
        new_ext = self.get_value("out_ext")
        out_img = splitext(orig_img)[0] + new_ext
        copy(os.path.join(chal_dir, orig_img), os.path.join(chal_dir, out_img))
        self.chal_file = out_img


class EditNotepad(GeneratedChallenge):
    yaml_tag = u'!edit_notepad'
    __doc__ = """
    Open in Notepad to find flag

    Config:

        orig_img - image file 
    """

    def gen(self, chal_dir):
        orig_img = self.get_value("orig_img")
        with open(os.path.join(chal_dir, orig_img), 'rb') as f:
            content = f.read()
            bytelist = bytearray(content)

        bytelist.extend(self.flag.encode('utf-8'))
        outputfile = 'edited_photo.png'

        with open(os.path.join(chal_dir, outputfile), 'wb') as image_file:
            image_file.write(bytelist)
        self.chal_file = outputfile


class SteghideChallenge(GeneratedChallenge):
    yaml_tag = u'!steghide'
    __doc__ = """
    Hide flag using steghide

    Config:

        text - addon text
        image - image to hide information in
        password - add password (optional)
        wordlist - wordlist for password cracking (optional)
    """

    def gen(self, chal_dir):
        text = self.get_value("text", required=False)
        image = self.get_value("image")
        password = self.get_value("password", required=False)
        password_command = "" if password is None else password
        word_list = self.get_value("word_list", required=False)

        words = [password]
        if word_list is not None:
            with open(join(chal_dir, word_list), "r") as w:
                words = [l.strip() for l in w.readlines()]

        assert password in words

        input_dir = "input"
        self.chal_file = 'chal' + splitext(image)[1]
        out_image_path = join(chal_dir, self.chal_file)
        full_text = text + " " + self.flag

        builder = DockerBuilder(
            base_img="ubuntu",
            input_dir=input_dir,
            included_files=[join(chal_dir, image)],
            output_files=[join(input_dir, image)],
            outdir=out_image_path
        )
        with builder as b:
            b.run('apt-get update', timeout=60)
            b.run('apt-get install -y steghide', timeout=60)
            b.run('cd input')
            b.run(f"echo '{full_text}' > flag.txt")
            b.run(
                f"steghide embed -p '{password_command}' -cf {image} -ef flag.txt", detach=False)


class ReverseAudio(GeneratedChallenge):
    yaml_tag = u'!rev_audio'
    __doc__ = """
    Reverse the audio to hear the flag

    Config:

        text - additional text to say
    """

    def ascii_to_phonetics(self, text):
        phonetics = {
            '&': 'and',
            '@': 'at',
            '#': 'hash',
            '$': 'dollar',
            '%': 'percent',
            '^': 'caret',
            '*': 'star',
            '+': 'plus',
            '=': 'equals',
            '\\': 'backslash',
            '/': 'slash',
            '<': 'less than',
            '>': 'greater than',
            '?': 'question mark',
            '!': 'exclamation point',
            '~': 'tilde',
            '`': 'backtick',
            '_': 'underscore',
            '-': 'dash',
            ':': 'colon',
            ';': 'semicolon',
            '"': 'double quote',
            "'": 'single quote',
            '|': 'pipe',
            '{': 'open brace',
            '}': 'close brace',
            '[': 'open bracket',
            ']': 'close bracket',
            '(': 'open parenthesis',
            ')': 'close parenthesis',
            ',': 'comma',
            '.': 'period',
            ' ': 'space',
        }
        if text in phonetics:
            return phonetics[text]
        else:
            return text

    def gen(self, chal_dir):
        out_file = join(chal_dir, 'chal.wav')
        phoentic_flag = [self.ascii_to_phonetics(i) for i in self.flag]
        spoken_flag = ",".join(phoentic_flag)
        full_text = self.get_value("text") + " " + spoken_flag
        text_to_wav(full_text, out_file, tld='co.uk')
        self.chal_file = out_file
        audio_data = AudioSegment.from_file(out_file)
        reversed_audio = audio_data.reverse()
        reversed_audio.export(out_file, format="wav")


class BinaryAudio(GeneratedChallenge):
    yaml_tag = u'!binary_audio'
    __doc__ = """
    Use binary audio to find flag

    Config:

        text - additional text to say
    """

    def gen(self, chal_dir):
        out_file = join(chal_dir, 'chal.wav')
        full_text = self.get_value("text") + " " + self.flag
        binary_full_text = binary_text(full_text)
        empty = AudioSegment.silent(duration=0)
        for c in binary_full_text:
            if c == ' ':
                empty += AudioSegment.silent(duration=200)
                continue
            elif c == '0':
                n = 1
            else:
                n = 10
            gen = Sine(200*n)
            sine = gen.to_audio_segment(duration=200).apply_gain(-3)
            sine = sine.fade_in(50).fade_out(100)
            empty += sine
        empty.export(out_file, format="wav")
        self.chal_file = out_file


class InvisibleChars(GeneratedChallenge):
    yaml_tag = u'!invisible_chars'
    __doc__ = """
    Use invisible characters to hide the flag

    Config:
    
        hide_text - text before invisible characters
        addl_text - additional text to say
    """

    def gen(self, chal_dir):
        out_file = join(chal_dir, 'chal.txt')
        hide_text = self.get_value("hide_text")
        addl_text = self.get_value("addl_text", required=False)
        full_text = self.flag
        if addl_text is not None:
            full_text += " " + addl_text
        binary_flag = binary_text(full_text)
        encoded_flag = [chr(6069) if c == '0' else chr(6158)
                        for c in binary_flag]
        chal_txt = hide_text + "".join(encoded_flag)
        self.display = chal_txt
        with open(out_file, 'w') as f:
            f.write(chal_txt)
        self.chal_file = out_file
